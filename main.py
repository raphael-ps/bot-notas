import openpyxl
import math
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

#COISAS LEGAIS PARA ADICIONAR
#UMA MANEIRA DE SABER SE A PESSOA SALVOU AS NOTAS ANTES DE IR PRA PROXIMA SALA
def continuidade(texto: str):
    print('-=-'*30)  
    while 1:
        print(texto+'\n')
        try:
            ask = int(input('--> '))
            if ask == 1:
                return False
            elif ask == 2:
                return True
            else:
                print("\nEscolha Inválida! Tente Novamente...")
                print('-=-'*30)
        except:
            print("\nEscolha Inválida! Tente Novamente...")
            print('-=-'*30)

def open_file():
    while 1:    

        excel_name = input("Digite o nome do arquivo: ").strip()
        if not '.' in excel_name:
            excel_name += '.xlsx'

        try:
            file_excel = openpyxl.load_workbook(filename=excel_name, data_only=True)
            print(f'Abrindo arquivo "{excel_name}"...')
            print("Arquivo aberto.")
            return file_excel

        except:
            
            print(f'\n[ERRO] - Arquivo "{excel_name}" não encontrado')

            if (continuidade('Deseja tentar reescrever o nome do arquivo?\n\n[1] - Sim\n[2] - Não')):
                return
            print('-=-'*30)

def excel_get_sheet(list_sheets):
    while 1:

        flag = True   
        print(f'{" ESSAS SÃO AS PAGINAS DO SEU ARQUIVO ":-^90}')  

        for index, sheet in enumerate(list_sheets):
            print(f"\t[{index+1:02}] - {sheet}")
        print(f"\t[{len(list_sheets)+1}] - ESCOLHER TODAS AS SALAS")

        picks = input("\nDIGITE O(S) NUMERO(s) DA(S) SALA(S) DESEJADA(S):\n-->").split()

        try:
            if int(picks[0]) == 11:
                lista = [n for n in range(0, len(list_sheets))]
                return lista

            else:
                for pos in range(0, len(picks)):
                    picks[pos] = int(picks[pos])-1

                    if picks[pos] > len(list_sheets) or picks[pos] < 0:
                        flag = False
                        break   

            if flag:
                return picks  

        except:
            print("Escolha(s) Inválida(s)! Tente Novamente...\n")


def get_data(work_sheet):
    data = dict()
    num_r = len(work_sheet['A'])
    agora = False
    for c in range(35, num_r):
        nome = work_sheet.cell(row=c, column=1).value
        media = work_sheet.cell(row = c, column = 2).value
        
        if nome != None and nome.strip() == "Aluno":
            agora = True
        elif (nome != None and media != None) and agora == True:
            media = math.ceil(media)
            nome = (nome[2:]).strip()       
            data.update({nome:media})
            print(nome, media)

    return data

def browser_init():
    navegador = webdriver.Chrome()
    navegador.maximize_window()
    navegador.get('https://professor.seduc.ce.gov.br/')
    navegador.find_element_by_class_name('btn-login').click()

    return navegador

def create_link(bimestre: int, sala: int):
    salas_id = {0:'685719',1:'685720',2:'685721',3:'685724',
                4:'685725',5:'685726',6:'685727',
                7:'691665',8:'685728',9:'685729'}
    
    link = 'https://professor.seduc.ce.gov.br/avaliacao/lista/?cd_escola=484&cd_turma='+salas_id[sala]+'&cd_disciplina=800&nr_periodo='+str(bimestre)
    return link

    
def passar_notas(navegador, dados_excel: dict):
    lista_alunos = WebDriverWait(navegador,10).until(EC.element_to_be_clickable((By.ID, "body-list")),'Não Achei3')
    lista_alunos = lista_alunos.find_elements(By.CSS_SELECTOR, '.div-card')

    for aluno_site in lista_alunos:
        nome_site = aluno_site.find_element(By.CSS_SELECTOR,'.text-success')
        nome_site = (nome_site.text).strip()

        for keys1, values1 in dados_excel.items():

            if keys1 == nome_site and values1 >= 6:

                print(f"\tENCONTRADO(A): {keys1:<45} -> Nota = {values1}")             
                values1 = str(values1)
                nota_div = aluno_site.find_element(By.CSS_SELECTOR, '.form-control')
                nota_div.clear()
                nota_div.send_keys(values1)
                dados_excel.pop(keys1)
                break

            elif keys1 == nome_site and values1 < 6:
                print(f'\tNÃO ADICIONADO(A): {keys1:<45} -> Nota = {values1} - NOTA MENOR QUE 6')
                break

def openfile_text():
    try:
        text_arquivo = open('Alunos sem nota ou nao encontrados.txt','w+')
        print("Arquivo '.txt' aberto.")
        return text_arquivo
    except FileNotFoundError:
        print("Arquivo texto não encontrado.")
    
def get_bimestre():
    while 1:
        print('-=-'*30)
        print("\t[1] - 1º BIMESTRE")
        print("\t[2] - 2º BIMESTRE")
        print("\t[3] - 3º BIMESTRE")
        print("\t[4] - 4º BIMESTRE")
        print("\t[5] - RECUPERAÇÃO FINAL\n")
        periodo = input('INFORME O BIMESTRE: ')

        try:
            periodo = int(periodo)

            if periodo >= 1 and periodo <=5:
                return periodo

            else:
                print("\nEssa opção não existe, tente novamente...")
        except:
            print("\nOpção Inválida! Tente Novamente...")
    
def create_avaliacao(navegador: webdriver, nome: str):
    element = WebDriverWait(navegador,10).until(EC.element_to_be_clickable((By.ID,'btn-add')),'Não Achei')
    element.click()
    name = WebDriverWait(navegador,10).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="form-cadastro"]/div/div[3]/div')),'Não Achei1')
    name = name.find_element(By.TAG_NAME, 'input')
    name.clear()
    name.send_keys(nome)
    navegador.find_element(By.ID, 'btn-save').click()
    element = WebDriverWait(navegador,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.btn-primary')),'Não Achei2')
    element.click()

def passTo_txt(dados: dict, text_file, current_sheet):
    if len(dados) == 0:
        print("Todos Os Alunos Foram Encontrados E Adicionados!")
        text_file.write("Todos Os Alunos Foram Encontrados E Adicionados!")
    else:
        print("Os Seguintes Alunos Não Foram Encontrados Ou Estão Com Nota ZERO:\n")

        for aluno, notas in dados.items():
            print(f"\t{current_sheet} - {aluno:45} - Nota: {notas}")
            text_file.write(current_sheet+' --> '+aluno.ljust(50) +' --> Nota: '+str(notas)+'\n')

        text_file.write('\n')
        print("\nLembre-se de adicionar a nota manualmente.")

def procura_avaliacao_existente(navegador: webdriver):
    try:
        WebDriverWait(navegador,10).until(EC.element_to_be_clickable((By.ID,'btn-add')),'Não Achei')
        template = navegador.find_element(By.CSS_SELECTOR, '.btn-primary')
        if not continuidade("JÁ EXISTE UMA AVALIAÇÃO, DESEJA EDITÁ-LA?\n[1] - SIM\n[2] - NÂO"):
            template.click()
            return True
        else:
            return False
    except:
        return False

print(f"\n{' BOT DAS NOTAS ':-^90}")
arquivo = open_file()

if not arquivo:
    print('Você escolheu sair...')
    print(f"{' THE END ':-^90}")    
    exit()

try:

    text_file = openfile_text()
    browser = browser_init()
    print('-=-'*30)
    input("AO TERMINAR DE FAZER O LOGIN NO SITE PRESSIONE 'ENTER'...")

    while 1:

        while 1:

            bimestre = get_bimestre()
            nome_avaliacao = (input("DIGITE O(S) NOME(S) DA(S) AVALIAÇÃO(ÕES): ").strip()).upper()
            sheets_list = arquivo.sheetnames
            sheets_pickeds = excel_get_sheet(sheets_list)

            print(" CONFIRME OS DADOS ".center(90,'-'))
            print(f"{'NOME DAS AVALIAÇÕES:':>20} {nome_avaliacao}\n{'BIMESTRE:':>20} {bimestre}\n{'SALA(S):':>20} [", end=' ')
            for index in sheets_pickeds:
                print(sheets_list[index], end=' ')
            print(']')
            if not continuidade("[1] - PARA CONFIRMAR\n[2] - PARA MUDAR OS DADOS"):
                break

        for cont in range(0, len(sheets_pickeds)):

            url = create_link(bimestre=bimestre, sala=sheets_pickeds[cont])
            current_sheet = sheets_list[sheets_pickeds[cont]]
            dados_excel = get_data(arquivo[current_sheet])

            browser.get(url)

            if not procura_avaliacao_existente(browser):
                create_avaliacao(navegador=browser, nome=nome_avaliacao)

            print('-=-'*30)
            passar_notas(browser, dados_excel)
            print('-=-'*30)
            passTo_txt(dados_excel, text_file, current_sheet)
            print('-=-'*30)

            del dados_excel, current_sheet, url
            input("Sala Atual Terminada... SALVE e Pressione ENTER para ir para a proxima...")

        print('-=-'*30)
        print("NÃO HÁ MAIS SALAS, ESQUECEU ALGUMA? ")

        if continuidade("[1] - SIM, Adicionar notas de outra sala\n[2] - NÃO, Encerrar programa."):
            break

finally:
    print('-=-'*30)
    print("Encerrando Script...")           
    print(f"{' THE END ':-^90}")

    arquivo.close()
    text_file.close()
    browser.quit() 
